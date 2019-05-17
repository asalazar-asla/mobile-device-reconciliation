import sys
import codecs
import csv
import xlrd
from collections import namedtuple
from copy import deepcopy
from openpyxl import load_workbook

def read_column(work_sheet, model_list, model_starting_row, column_pos):
    # Purpose: Read the first column of Device Data sheet
    # Input: work_sheet = worksheet from the Device Data XLSX file
    # Output: Retrieve the device's model and return to an empty list

    for row in work_sheet.iter_rows(column_pos.format(model_starting_row, work_sheet.max_row)):
        for cell in row:
            model_list.append(cell.value)

    return model_list

def get_device_names(model_list, device_dict, open_file_func):
    # Purpose: Compare list to the Model column in the CSV file
    # Input: List of internal model devices from Device Data XLSX file
    # Output: Dictionary entailing marketing and brand Android device name

    device_table = open_file_func()

    for row in device_table:
        if row[3] in model_list:
            key = row[3]
            device_name = f"{row[0]} {row[1]}"
            device_dict[key] = device_name

    return device_dict
    
def open_android_devices():
    return csv.reader(codecs.open(r"supported_devices.csv", 'rU', 'utf-16'))

def open_ios_devices():
    device_table = []
    with open(r"apple_devices.csv") as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            row[3] = f"{row[2]},{row[3]}"
            device_table.append(row)
    return device_table

def get_tier(device_dict, pivot_workbook, pivot_sheet_name):
    # Purpose: Compare android device name dictionary to XLSX file to retireve the counterpoint tier number
    # Input: device_dict = Dictionary of internal models as keys and
    # its values as marketing and brand Android device name 
    # Output: Update the dictionary to chnage it's values to tuples containing the Android
    # device name and its tier number or undefined if not found in the XLSX file

    pivot = xlrd.open_workbook(
        pivot_workbook, encoding_override="utf-8")
    pivot_android = pivot.sheet_by_name(pivot_sheet_name)

    pivot_table = []

    for devices in range(5,5651):
        pivot_table.append(str(pivot_android.cell(devices, 0).value))

    TierPair = namedtuple('TierPair', ['device_name', 'tier'])

    copyofDict = dict(device_dict)

    for index, (internal_name, device_name) in enumerate(copyofDict.items()):
        if device_name in pivot_table:
            temp = pivot_table.index(device_name)+1
            tier = int(float(pivot_table[temp]))
            device_dict[internal_name] = TierPair(device_name, tier)
        else:
            device_dict[internal_name] = TierPair(device_name, "undefined")

    return device_dict

def write_to_columns(work_sheet, col_num, device_dict, model_starting_row, column_pos):
    # Purpose: Write content from list to the designated column in the worksheet
    # Input: work_sheet: worksheet from the inputted XLSX file
    #        col_num: int value of designated column where it will be written
    #        device_dict: Dictionary containing the contents that will be inputted in
    # Output: List write-in to the cells in the designated column

    for row in work_sheet.iter_rows(column_pos.format(model_starting_row, work_sheet.max_row)):
        for cell in row:
            if cell.value in device_dict.keys():
                write_device_name = work_sheet.cell(row=cell.row, column=col_num)
                write_tier = work_sheet.cell(row=cell.row, column=col_num+1)
                write_device_name.value = device_dict[cell.value].device_name
                write_tier.value = device_dict[cell.value].tier
