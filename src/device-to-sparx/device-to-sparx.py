import sys
import codecs
import csv
import xlrd
import argparse
from openpyxl import load_workbook

def read_column(work_sheet, input_list, starting_row, column_pos):
    for row in work_sheet.iter_rows(starting_row, work_sheet.max_row, column_pos, column_pos):
        for cell in row:
            input_list.append(str(cell.value))

    return input_list


def get_device_model(device_list, model_dict, open_file_func):
    device_table = open_file_func()

    for row in device_table:
        name = f"{row[0]} {row[1]}"
        if name in device_list:
            key = name
            device_model = row[3]
            model_dict[key] = device_model
    
    return model_dict

def open_android_devices():
    return csv.reader(codecs.open(r"supported_devices.csv", 'rU', 'utf-16'))

def open_ios_devices():
    device_table = []
    with open(r"apple_devices.csv") as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            row[3] = f"{row[2]},{row[3]}"
            device_table.append(row)

    return device_table

def get_family(work_sheet, device_list, family_dict):

    for row in work_sheet.iter_rows(2, work_sheet.max_row, 10, 10):
        for cell in row:
            if cell.value in device_list:
                get_fam = str(work_sheet.cell(row=cell.row, column=27).value)

                family_dict[cell.value] = get_fam
    
    return family_dict


def write_to_column(work_sheet, write_column, model_dict, starting_row, column_pos):
    for row in work_sheet.iter_rows(starting_row, work_sheet.max_row, column_pos, column_pos):
        for cell in row:
            if cell.value in model_dict:
                write_device_model = work_sheet.cell(
                    row=cell.row, column=write_column)

                write_device_model.value = model_dict[cell.value]

def write_family(work_sheet, write_column, family_dict, starting_row, column_pos):
    for row in work_sheet.iter_rows(starting_row, work_sheet.max_row, column_pos, column_pos):
        for cell in row:
            if cell.value in family_dict:
                write_family = work_sheet.cell(
                    row=cell.row, column=write_column+1)

                write_family.value = family_dict[cell.value]

def main():
    # load data
    device_map = load_workbook(r"1_GB_devices_metrics.xlsx")
    android_sheet = device_map.get_sheet_by_name("Android")
    apple_sheet = device_map.get_sheet_by_name("iOS")
    device_coverage = load_workbook(r"Device Tiering and Coverage_Sept2018.xlsx")
    raw_sheet = device_coverage.get_sheet_by_name("Raw Data Sept2018")

    device_list = []
    model_dict = {}
    family_dict = {}

    starting_row = 2

    device_list = read_column(android_sheet, device_list,
                              starting_row, 1)
    model_dict = get_device_model(device_list, model_dict, open_android_devices)
    family_dict = get_family(raw_sheet, device_list, family_dict)

    write_to_column(android_sheet, 2, model_dict, starting_row, 1)
    write_family(android_sheet, 2, family_dict, starting_row, 1)

    device_list = read_column(apple_sheet, device_list, starting_row, 1)
    model_dict = get_device_model(device_list, model_dict, open_ios_devices)
    family_dict = get_family(raw_sheet, device_list, family_dict)

    write_to_column(apple_sheet, 2, model_dict, starting_row, 1)
    write_family(apple_sheet, 2, family_dict, starting_row, 1)

    device_map.save(r"1_GB_Devices_metrics.xlsx")
    print(f"Mapping complete")

if __name__ == "__main__":

    main()